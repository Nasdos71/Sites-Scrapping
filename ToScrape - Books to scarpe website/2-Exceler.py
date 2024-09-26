import sys
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Fill
from os import curdir as dir, listdir as lis, makedirs, path, getcwd

wb = Workbook()
ws = wb.active
ws['A1'] = 'Category'
ws['B1'] = 'Name'
ws['C1'] = 'Price'
ws['D1'] = 'Status'
a = ws['A1']
b = ws['B1']
c = ws['C1']
d = ws['D1']

a.font = Font(color='0099CCFF', bold=True)
b.font = Font(color ='0099CCFF', bold=True)
c.font = Font(color ='0099CCFF', bold=True)
d.font = Font(color ='0099CCFF', bold=True)






dirt = getcwd()
output_dir = path.join(dirt, "Data Excel")
makedirs(output_dir, exist_ok=True)
Path = path.join(dirt, 'Books')
n = p = s = None
row  = 2
for t in lis(Path):
    if t.endswith('.txt'):
        pp = path.join(Path, t)
        with open(pp, 'r') as f:
            catg = f.readline().split(":")[1]
            reader = f.readlines()

            for i in range(1, len(reader)):
                sp = reader[i]
                if sp.startswith("Name :"):
                    n = sp.split(':')[1].strip()
                    if len(sp.split(':')) > 2:
                        n += sp.split(':')[2].strip()

                if sp.startswith("Price :"):
                    p = sp.split(':')[1].strip()
                    

                if sp.startswith("Status :"):
                    s = sp.split(':')[1].strip()
                    
                if n  and p  and s :
                    ws[f'A{row}'] = catg
                    ws[f'B{row}'] = n
                    ws[f'C{row}'] = p
                    ws[f'D{row}'] = s
                    row += 1 
                    n = p = s = None
                    continue

wb.save(path.join(output_dir, "Data.xlsx"))