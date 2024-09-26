import random
from openpyxl import load_workbook as l_w
from openpyxl.styles import Font, Fill
import os


fs = ['Nasef', 'Ziad', 'Iyad', 'Omar', 'Walied', 'Mohamed', 'Ahmad', 'Ismail', 'Youesef', 'Abdelrahman', 'Abdallah', 'Hamada', 'Mahmoud', 'Hisham', 'Salem', 'Saleh']
ls = ['Tawfik', 'El-Shimy', 'Hegab', 'Omara', 'Farouk', 'Assad', 'Kooud', 'Matbooly', 'El-Hassan', 'Metwaly', 'El-Lahham', 'Mariouty', 'Maghraby', 'El-Beheiry']

auth = set() 


while len(auth) < 50:
    first_name = random.choice(fs)
    last_name = random.choice(ls)
    full_name = f"{first_name} {last_name}"
    auth.add(full_name)  


auth = list(auth)

boom = os.getcwd()
file = os.path.join(boom, 'Data Excel', 'Data.xlsx' )
wb = l_w(file)

ws = wb.active

ws['E1'] = 'Quantity'
ws['F1']= 'Author'
a = ws['E1']
b = ws['F1']

a.font = Font(color ='0099CCFF', bold=True)
b.font = Font(color ='0099CCFF', bold=True)

for i in range(2, 1002):
    ws[F"E{i}"] = random.randint(1,200)
    ws[f"F{i}"] = random.choice(auth)

wb.save(file)


